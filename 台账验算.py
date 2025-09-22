import openpyxl
from openpyxl.styles import PatternFill
from datetime import datetime
import os
import sys
import re
import logging
import tkinter as tk
from tkinter import filedialog, messagebox
import threading

# -------------------- 日志设置 --------------------
logger = logging.getLogger()
logger.setLevel(logging.INFO)

# 创建文件处理器并设置编码为utf-8
file_handler = logging.FileHandler('execution.log', encoding='utf-8')
file_handler.setLevel(logging.INFO)

# 创建日志格式
formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
file_handler.setFormatter(formatter)

# 将处理器添加到日志记录器
logger.addHandler(file_handler)

# -------------------- 辅助函数 --------------------

def clean_amount(value):
    """
    清理和转换金额数据。
    移除非数字字符（如逗号、货币符号等），然后尝试转换为float。
    如果转换失败，返回None。
    """
    if isinstance(value, (int, float)):
        return value
    if isinstance(value, str):
        # 移除所有非数字和非小数点的字符
        cleaned = re.sub(r'[^\d.-]', '', value)
        try:
            return float(cleaned)
        except ValueError:
            return None
    return None

def center_window(win, width, height):
    """
    将窗口居中显示在屏幕上。
    """
    screen_width = win.winfo_screenwidth()
    screen_height = win.winfo_screenheight()
    x = int((screen_width / 2) - (width / 2))
    y = int((screen_height / 2) - (height / 2))
    win.geometry(f"{width}x{height}+{x}+{y}")

# -------------------- 功能一：更新还款状态 --------------------

def update_repayment_status(ws_order, ws_flow, current_date):
    logging.info("开始执行功能一：更新还款状态...")
    try:
        # 构建付款信息字典
        payment_info = {}
        last_row_flow = ws_flow.max_row

        for flow_row in range(2, last_row_flow + 1):
            order_num = ws_flow.cell(row=flow_row, column=2).value  # 列B
            term = ws_flow.cell(row=flow_row, column=10).value  # 列J
            payment_date = ws_flow.cell(row=flow_row, column=1).value  # 列A
            payment_type = ws_flow.cell(row=flow_row, column=11).value  # 列K

            if payment_type in ["租金", "尾款"] and isinstance(payment_date, datetime):
                key = f"{order_num}_{term}"
                if key not in payment_info:
                    payment_info[key] = payment_date.date()

        logging.info(f"收集到的付款信息数量: {len(payment_info)}")

        last_row_order = ws_order.max_row

        # 获取期数列表（假设期数在第4行，列14到23，即N到W列）
        term_row = 4
        term_columns = range(14, 24)  # 列N到W

        # 定义颜色填充
        color_overdue = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # 逾期还款 - 浅黄色
        color_early = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")    # 提前还款 - 浅蓝色
        color_on_time = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # 按时还款 - 浅绿色
        color_bill_day = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid") # 账单日 - 橙色
        color_overdue_not_paid = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid") # 逾期未还 - 浅红色
        color_invalid = PatternFill(start_color="DEDEDE", end_color="DEDEDE", fill_type="solid") # 无效或空白 - 灰色

        for order_row in range(5, last_row_order + 1):
            order_num = ws_order.cell(row=order_row, column=2).value  # 列B
            if not order_num:
                continue  # 跳过空订单号

            for col in term_columns:
                term = ws_order.cell(row=term_row, column=col).value
                due_date = ws_order.cell(row=order_row, column=col).value

                if isinstance(due_date, datetime):
                    due_date = due_date.date()
                    key = f"{order_num}_{term}"

                    if key in payment_info:
                        payment_date = payment_info[key]
                        if payment_date > due_date:
                            fill = color_overdue  # 逾期还款
                        elif payment_date < due_date:
                            fill = color_early    # 提前还款
                        else:
                            fill = color_on_time   # 按时还款
                    else:
                        if current_date == due_date:
                            fill = color_bill_day  # 账单日
                        elif current_date > due_date:
                            fill = color_overdue_not_paid  # 逾期未还
                        else:
                            fill = None  # 未到还款日期，不变色

                    if fill:
                        ws_order.cell(row=order_row, column=col).fill = fill
                    else:
                        ws_order.cell(row=order_row, column=col).fill = PatternFill(fill_type=None)

                else:
                    # 无效或空白的还款日期
                    ws_order.cell(row=order_row, column=col).fill = color_invalid

        logging.info("功能一：更新还款状态完成。")
    except Exception as e:
        logging.error(f"功能一执行失败: {e}")
        raise

# -------------------- 功能二：计算当前应收款和成本 --------------------

def calculate_receivables_and_costs(ws_order, ws_flow):
    logging.info("开始执行功能二：计算当前应收款和成本...")
    try:
        receivables_dict = {}
        costs_dict = {}
        last_row_flow = ws_flow.max_row

        for flow_row in range(2, last_row_flow + 1):
            order_id = ws_flow.cell(row=flow_row, column=2).value  # 列B
            payment_type = ws_flow.cell(row=flow_row, column=11).value  # 列K
            transaction_amount_raw = ws_flow.cell(row=flow_row, column=9).value  # 列I
            transaction_amount = clean_amount(transaction_amount_raw)

            if not order_id:
                continue  # 跳过空订单ID

            if transaction_amount is None:
                logging.warning(f"第 {flow_row} 行的交易金额 '{transaction_amount_raw}' 无法转换为数值。该行将被跳过。")
                continue  # 跳过无法转换的金额

            # 应收款
            if payment_type in ["尾款", "租金", "首付款"]:
                receivables_dict[order_id] = receivables_dict.get(order_id, 0) + transaction_amount

            # 成本
            if payment_type in ["放款", "供应商利润"]:
                costs_dict[order_id] = costs_dict.get(order_id, 0) + abs(transaction_amount)

        logging.info(f"计算得到的应收款订单数量: {len(receivables_dict)}")
        logging.info(f"计算得到的成本订单数量: {len(costs_dict)}")

        last_row_order = ws_order.max_row

        for order_row in range(2, last_row_order + 1):
            order_id = ws_order.cell(row=order_row, column=2).value  # 列B
            if not order_id:
                continue  # 跳过空订单ID

            # 更新应收款 - 列L (12)
            initial_receivables_raw = ws_order.cell(row=order_row, column=11).value  # 列K
            initial_receivables = clean_amount(initial_receivables_raw)
            if initial_receivables is not None:
                receivables_paid = receivables_dict.get(order_id, 0)
                current_receivables = initial_receivables - receivables_paid
                ws_order.cell(row=order_row, column=12).value = current_receivables  # 列L
            else:
                logging.warning(f"订单管理表第 {order_row} 行的初始应收款 '{initial_receivables_raw}' 无法转换为数值。应收款列将直接复制初始值。")
                ws_order.cell(row=order_row, column=12).value = ws_order.cell(row=order_row, column=11).value  # 列L等于列K

            # 更新成本 - 列Y (25)
            total_cost = costs_dict.get(order_id, 0)
            ws_order.cell(row=order_row, column=25).value = total_cost  # 列Y

        logging.info("功能二：计算当前应收款和成本完成。")
    except Exception as e:
        logging.error(f"功能二执行失败: {e}")
        raise

# -------------------- 功能三：填充交易金额 --------------------

def fill_transaction_amounts(ws_order, ws_flow):
    logging.info("开始执行功能三：填充交易金额...")
    try:
        transaction_dict = {}
        last_row_flow = ws_flow.max_row

        for flow_row in range(2, last_row_flow + 1):
            order_id = ws_flow.cell(row=flow_row, column=2).value  # 列B
            term = ws_flow.cell(row=flow_row, column=10).value  # 列J
            transaction_amount_raw = ws_flow.cell(row=flow_row, column=9).value  # 列I
            transaction_amount = clean_amount(transaction_amount_raw)

            if not order_id:
                continue  # 跳过空订单ID

            if term == "第一期" and transaction_amount is not None:
                if order_id not in transaction_dict:
                    transaction_dict[order_id] = transaction_amount
                else:
                    # 如果订单号已存在，累加交易金额
                    transaction_dict[order_id] += transaction_amount

        logging.info(f"收集到的第一期交易信息数量: {len(transaction_dict)}")

        last_row_order = ws_order.max_row

        for order_row in range(2, last_row_order + 1):
            order_id = ws_order.cell(row=order_row, column=2).value  # 列B
            if not order_id:
                continue  # 跳过空订单ID

            if order_id in transaction_dict:
                ws_order.cell(row=order_row, column=24).value = transaction_dict[order_id]  # 列X
            else:
                ws_order.cell(row=order_row, column=24).value = "未找到匹配"  # 列X

        logging.info("功能三：填充交易金额完成。")
    except Exception as e:
        logging.error(f"功能三执行失败: {e}")
        raise

# -------------------- 功能四：更新余额并复制订单详情 --------------------

def update_balances_and_copy_order_details(ws_order, ws_flow, ws_tools):
    logging.info("开始执行功能四：更新余额并复制订单详情...")
    try:
        # 设置条件范围
        rng_criteria_available = set()
        rng_criteria_withdraw = set()

        # 读取工具表中的B6:B15
        for row in ws_tools.iter_rows(min_row=6, max_row=15, min_col=2, max_col=2):
            cell_value = row[0].value
            if cell_value:
                rng_criteria_available.add(cell_value)

        # 读取工具表中的B2:B5
        for row in ws_tools.iter_rows(min_row=2, max_row=5, min_col=2, max_col=2):
            cell_value = row[0].value
            if cell_value:
                rng_criteria_withdraw.add(cell_value)

        logging.info(f"可用余额条件数量: {len(rng_criteria_available)}")
        logging.info(f"提现余额条件数量: {len(rng_criteria_withdraw)}")

        # 获取订单管理工作表的最后一行
        order_last_row = ws_order.max_row

        # 获取资金流水账工作表的最后一行
        ledger_last_row = ws_flow.max_row

        # 构建订单管理字典，键为订单号，值为C到H列的值
        order_dict = {}
        for order_row in range(2, order_last_row + 1):
            order_number = ws_order.cell(row=order_row, column=2).value  # 列B
            if not order_number:
                continue  # 跳过空订单号
            # 获取C到H列的值
            order_details = [ws_order.cell(row=order_row, column=col).value for col in range(3, 9)]  # 列C到H
            order_dict[order_number] = order_details

        logging.info(f"订单管理表中订单数量: {len(order_dict)}")

        # 初始化余额
        available_balance = 0.0
        withdraw_balance = 0.0

        # 定义颜色填充（如果需要，可根据需求添加）

        for ledger_row in range(5, ledger_last_row + 1):
            # 检查I列是否有数值，以便更新余额
            transaction_amount_raw = ws_flow.cell(row=ledger_row, column=9).value  # 列I
            transaction_amount = clean_amount(transaction_amount_raw)  # 列I

            if transaction_amount is not None:
                transaction_type = ws_flow.cell(row=ledger_row, column=12).value  # 列L

                # 更新可用余额
                if transaction_type in rng_criteria_available:
                    available_balance += transaction_amount
                ws_flow.cell(row=ledger_row, column=14).value = available_balance  # 列N

                # 更新提现余额
                if transaction_type == "复投卡":
                    withdraw_balance -= transaction_amount
                elif transaction_type in rng_criteria_withdraw:
                    withdraw_balance += transaction_amount
                ws_flow.cell(row=ledger_row, column=15).value = withdraw_balance  # 列O
            else:
                logging.warning(f"资金流水账表第 {ledger_row} 行的交易金额 '{transaction_amount_raw}' 无法转换为数值。余额将不更新。")

            # 复制订单详情
            order_number = ws_flow.cell(row=ledger_row, column=2).value  # 列B
            if order_number:
                if order_number in order_dict:
                    order_details = order_dict[order_number]
                    for idx, value in enumerate(order_details, start=3):  # 列C到H
                        ws_flow.cell(row=ledger_row, column=idx).value = value
                else:
                    logging.warning(f"订单号 {order_number} 在订单管理表中未找到匹配。")

        logging.info("功能四：更新余额并复制订单详情完成。")
    except Exception as e:
        logging.error(f"功能四执行失败: {e}")
        raise

# -------------------- GUI 应用 --------------------

def run_processing(file_path):
    try:
        # 加载Excel工作簿
        wb = openpyxl.load_workbook(file_path, keep_vba=True)
        ws_order = wb["订单管理"]
        ws_flow = wb["资金流水账"]
        ws_tools = wb["工具表"]
    except Exception as e:
        logging.error(f"加载Excel文件时出错: {e}")
        messagebox.showerror("错误", f"加载Excel文件时出错: {e}")
        return

    # 获取当前日期
    current_date = datetime.now().date()

    try:
        # 执行功能一
        update_repayment_status(ws_order, ws_flow, current_date)

        # 执行功能二
        calculate_receivables_and_costs(ws_order, ws_flow)

        # 执行功能三
        fill_transaction_amounts(ws_order, ws_flow)

        # 执行功能四
        update_balances_and_copy_order_details(ws_order, ws_flow, ws_tools)

        # 保存工作簿
        wb.save(file_path)
        logging.info(f"所有功能执行完成，Excel文件已保存：{file_path}")
        messagebox.showinfo("完成", "所有功能执行完成，Excel文件已保存。")
    except Exception as e:
        logging.error(f"执行过程中出错: {e}")
        messagebox.showerror("错误", f"执行过程中出错: {e}")

def select_file():
    file_path = filedialog.askopenfilename(
        title="选择Excel文件",
        filetypes=[("Excel Files", "*.xlsm *.xlsx *.xls")]
    )
    if file_path:
        entry_file_path.delete(0, tk.END)
        entry_file_path.insert(0, file_path)

def execute_processing():
    file_path = entry_file_path.get()
    if not file_path:
        messagebox.showwarning("警告", "请先选择一个Excel文件。")
        return
    if not os.path.isfile(file_path):
        messagebox.showerror("错误", "所选文件不存在。")
        return
    # 显示“正在验算中”的弹窗
    processing_window = tk.Toplevel(root)
    processing_window.title("正在验算中")
    processing_window.geometry("300x100")
    center_window(processing_window, 300, 100)
    processing_label = tk.Label(processing_window, text="正在验算中，请稍候...")
    processing_label.pack(padx=20, pady=20)
    # 禁用主窗口的按钮
    btn_execute.config(state='disabled')
    btn_view_log.config(state='disabled')
    btn_select.config(state='disabled')
    btn_exit.config(state='disabled')
    root.update_idletasks()
    # 运行处理任务在新线程
    def task():
        run_processing(file_path)
        # 关闭“正在验算中”弹窗
        processing_window.destroy()
        # 重新启用主窗口的按钮
        btn_execute.config(state='normal')
        btn_view_log.config(state='normal')
        btn_select.config(state='normal')
        btn_exit.config(state='normal')

    threading.Thread(target=task).start()

def view_log():
    log_path = 'execution.log'
    if os.path.isfile(log_path):
        try:
            with open(log_path, 'r', encoding='utf-8', errors='replace') as log_file:
                log_content = log_file.read()
            # 创建一个新的窗口显示日志
            log_window = tk.Toplevel(root)
            log_window.title("日志内容")
            log_window.geometry("800x600")
            center_window(log_window, 800, 600)
            log_text = tk.Text(log_window, wrap='word', width=100, height=30)
            log_text.pack(expand=True, fill='both')
            log_text.insert(tk.END, log_content)
            log_text.config(state='disabled')
        except Exception as e:
            messagebox.showerror("错误", f"无法读取日志文件: {e}")
    else:
        messagebox.showinfo("信息", "日志文件尚不存在。")

def execute_with_processing():
    # Run in a separate thread to avoid blocking the GUI
    execute_processing()

def main_gui():
    global root, entry_file_path, btn_execute, btn_view_log, btn_select, btn_exit
    root = tk.Tk()
    root.title("Excel 数据处理工具")

    # 设置窗口大小和位置
    window_width = 600
    window_height = 200
    center_window(root, window_width, window_height)

    # 文件选择部分
    frame_file = tk.Frame(root, padx=10, pady=10)
    frame_file.pack(fill='x')

    btn_select = tk.Button(frame_file, text="选择Excel文件", command=select_file)
    btn_select.pack(side='left')

    entry_file_path = tk.Entry(frame_file, width=60)
    entry_file_path.pack(side='left', padx=5)

    # 执行按钮
    frame_execute = tk.Frame(root, padx=10, pady=10)
    frame_execute.pack(fill='x')

    btn_execute = tk.Button(frame_execute, text="一键执行", command=execute_with_processing, bg='green', fg='white')
    btn_execute.pack(side='left')

    # 查看日志按钮
    btn_view_log = tk.Button(frame_execute, text="查看日志", command=view_log)
    btn_view_log.pack(side='left', padx=5)

    # 退出按钮
    btn_exit = tk.Button(frame_execute, text="退出", command=root.quit, fg='red')
    btn_exit.pack(side='right')

    root.mainloop()

if __name__ == "__main__":
    main_gui()
